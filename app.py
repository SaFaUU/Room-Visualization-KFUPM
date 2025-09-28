import io
from datetime import time as dtime, datetime, timedelta
from collections import defaultdict
import math

import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

st.set_page_config(page_title="Room Assignment & Conflict Resolver", layout="wide")

# -------------------------------
# Helpers
# -------------------------------
HEADER_TOKENS = {"subject", "number", "title", "section", "section_act", "sec", "days", "start", "end", "bldg", "room"}
DAY_ORDER = ["U","M","T","W","R","F","S"]  # Sunday-first
DAY_TO_OFFSET = {d:i for i,d in enumerate(DAY_ORDER)}
WEEK_ANCHOR = datetime(2025, 1, 5)  # A Sunday

def norm_str(x):
    if pd.isna(x): return ""
    return str(x).strip()

def parse_days(d):
    d = norm_str(d).upper().replace(" ", "")
    valid = set(DAY_ORDER)
    if d == "" or d == "TBA":
        return []
    return [ch for ch in d if ch in valid]

def parse_time(t):
    if pd.isna(t): return None
    if isinstance(t, dtime):
        return t.hour*60 + t.minute
    # Some spreadsheets use 1400, 1640
    try:
        s = str(int(float(t))).zfill(4)
        hh, mm = int(s[:2]), int(s[2:])
        return hh*60 + mm
    except Exception:
        s = str(t).strip()
        if ":" in s:
            try:
                hh, mm = s.split(":")[:2]
                return int(hh)*60 + int(mm)
            except Exception:
                return None
        return None

def time_str(minutes):
    if minutes is None: return ""
    hh = minutes // 60
    mm = minutes % 60
    return f"{hh:02d}:{mm:02d}"

def minutes_to_dt(day_char, minutes):
    base = WEEK_ANCHOR + timedelta(days=DAY_TO_OFFSET.get(day_char, 0))
    return base.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=int(minutes))

def overlaps(a_start, a_end, b_start, b_end):
    if None in (a_start, a_end, b_start, b_end): 
        return False
    return (a_start < b_end) and (b_start < a_end)

def detect_header_row(df, search_rows=10):
    best_row, best_score = 0, -1
    for i in range(min(search_rows, len(df))):
        row = df.iloc[i].astype(str).str.strip().str.lower()
        score = sum(1 for cell in row for tok in HEADER_TOKENS if tok in str(cell))
        if score > best_score:
            best_score, best_row = score, i
    return best_row

def normalize_columns(df):
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    return df

def build_rooms(df):
    # Build list of unique rooms from assigned_bldg/assigned_room if present, else bldg/room
    if "assigned_bldg" in df.columns and "assigned_room" in df.columns:
        cols = ["assigned_bldg","assigned_room"]
    else:
        cols = ["bldg","room"]
    tmp = df[cols].dropna(how="any")
    if tmp.empty:
        return []
    rooms = (
        tmp.astype(str)
        .applymap(lambda x: x.strip())
        .drop_duplicates()
        .to_records(index=False)
        .tolist()
    )
    return rooms

def rebuild_occupancy(df):
    from ast import literal_eval
    occ = defaultdict(list)
    for idx, row in df.iterrows():
        b = norm_str(row.get("assigned_bldg"))
        r = norm_str(row.get("assigned_room"))
        days_list = row.get("days_list", [])
        if isinstance(days_list, str):
            try:
                days_list = literal_eval(days_list)
            except Exception:
                days_list = []
        if b == "" or r == "" or len(days_list) == 0 or row.get("start_min") is None or row.get("end_min") is None:
            continue
        for d in days_list:
            occ[(d, b, r)].append(idx)
    return occ

def find_free_room(occupancy, df, rooms, day, start_min, end_min, skip_idx=None):
    for (b, r) in rooms:
        busy = False
        for idx2 in occupancy.get((day, b, r), []):
            if skip_idx is not None and idx2 == skip_idx: 
                continue
            row2 = df.loc[idx2]
            if overlaps(start_min, end_min, row2["start_min"], row2["end_min"]):
                busy = True
                break
        if not busy:
            return b, r
    return None, None

def list_free_rooms_same_time(occupancy, df, rooms, day, start_min, end_min, limit=10, exclude=None):
    if exclude is None: exclude = set()
    alternatives = []
    for (b, r) in rooms:
        if (b, r) in exclude: 
            continue
        busy = False
        for idx2 in occupancy.get((day, b, r), []):
            row2 = df.loc[idx2]
            if overlaps(start_min, end_min, row2["start_min"], row2["end_min"]):
                busy = True
                break
        if not busy:
            alternatives.append((b, r))
        if len(alternatives) >= limit: 
            break
    return alternatives

def free_slots_for_room(df, day, bldg, room, duration_min, day_start=8*60, day_end=20*60, limit=10):
    from ast import literal_eval
    entries = []
    for _, row in df.iterrows():
        if bldg == norm_str(row.get("assigned_bldg")) and room == norm_str(row.get("assigned_room")):
            days_list = row.get("days_list", [])
            if isinstance(days_list, str):
                try: days_list = literal_eval(days_list)
                except Exception: days_list = []
            if day in days_list and row.get("start_min") is not None and row.get("end_min") is not None:
                entries.append((int(row["start_min"]), int(row["end_min"])))
    entries.sort()
    gaps = []
    prev = int(day_start)
    for (s,e) in entries:
        if s - prev >= duration_min:
            gaps.append((prev, s))
        prev = max(prev, e)
    if day_end - prev >= duration_min:
        gaps.append((prev, int(day_end)))
    res = []
    for (s,e) in gaps:
        if e - s >= duration_min:
            res.append((s, s+duration_min))
        if len(res) >= limit:
            break
    return res

def list_free_slots_across_rooms(df, day, duration_min, day_start=8*60, day_end=20*60, limit=20, exclude_room=None):
    rooms = build_rooms(df)
    suggestions = []
    for (b, r) in rooms:
        if exclude_room and (b, r) == exclude_room:
            continue
        slots = free_slots_for_room(df, day, norm_str(b), norm_str(r), duration_min, day_start, day_end, limit=3)
        for (s,e) in slots:
            suggestions.append({"room": f"{b}-{r}", "day": day, "start": s, "end": e})
    suggestions.sort(key=lambda x: (x["start"], x["room"]))
    return suggestions[:limit]

def run_auto_resolve(df, prioritize_labs=True, max_rounds=3):
    rooms = build_rooms(df)
    occupancy = rebuild_occupancy(df)
    conflict_rows = []
    for _ in range(max_rounds):
        changed = False
        for key, idxs in list(occupancy.items()):
            day, bldg, room = key
            idxs_sorted = sorted(idxs, key=lambda i: (df.loc[i, "start_min"] if pd.notna(df.loc[i, "start_min"]) else 0))
            for i in range(len(idxs_sorted)):
                for j in range(i+1, len(idxs_sorted)):
                    a, b = idxs_sorted[i], idxs_sorted[j]
                    ra, rb = df.loc[a], df.loc[b]
                    if overlaps(ra["start_min"], ra["end_min"], rb["start_min"], rb["end_min"]):
                        keep_idx, move_idx = a, b
                        if prioritize_labs and rb["is_lab"] and not ra["is_lab"]:
                            keep_idx, move_idx = b, a
                        elif prioritize_labs and rb["is_lab"] and ra["is_lab"]:
                            conflict_rows.append({"idx": a, "day": day, "assigned_bldg": bldg, "assigned_room": room, 
                                                  "type": "hard", "reason": "Overlap between two labs",
                                                  "other_course": df.loc[b, "course_id"]})
                            conflict_rows.append({"idx": b, "day": day, "assigned_bldg": bldg, "assigned_room": room, 
                                                  "type": "hard", "reason": "Overlap between two labs",
                                                  "other_course": df.loc[a, "course_id"]})
                            continue
                        rm_b, rm_r = find_free_room(occupancy, df, rooms, day, df.loc[move_idx, "start_min"], df.loc[move_idx, "end_min"], move_idx)
                        if rm_b is not None:
                            old_b, old_r = norm_str(df.loc[move_idx, "assigned_bldg"]), norm_str(df.loc[move_idx, "assigned_room"])
                            df.at[move_idx, "assigned_bldg"] = rm_b
                            df.at[move_idx, "assigned_room"] = rm_r
                            occupancy = rebuild_occupancy(df)
                            conflict_rows.append({"idx": move_idx, "day": day, "assigned_bldg": rm_b, "assigned_room": rm_r, 
                                                  "type": "auto-resolved", 
                                                  "reason": f"Moved from {old_b}-{old_r}; prioritized lab" if prioritize_labs else f"Moved from {old_b}-{old_r}",
                                                  "other_course": df.loc[keep_idx, "course_id"]})
                            changed = True
                        else:
                            alts = list_free_rooms_same_time(occupancy, df, rooms, day, df.loc[move_idx, "start_min"], df.loc[move_idx, "end_min"],
                                                   limit=6, exclude={(norm_str(df.loc[move_idx,"assigned_bldg"]), norm_str(df.loc[move_idx,"assigned_room"]))})
                            conflict_rows.append({"idx": move_idx, "day": day, "assigned_bldg": df.loc[move_idx, "assigned_bldg"], "assigned_room": df.loc[move_idx, "assigned_room"], 
                                                  "type": "needs-manual", 
                                                  "reason": "No free room at same time",
                                                  "suggestions": "; ".join([f"{b}-{r}" for (b,r) in alts]) if alts else "(no same-time alternatives)",
                                                  "other_course": df.loc[keep_idx, "course_id"]})
        if not changed:
            break

    df["conflict_flag"] = False
    df["conflict_note"] = ""
    for c in conflict_rows:
        df.at[c["idx"], "conflict_flag"] = True
        parts = [c["type"], c.get("reason","")]
        if c.get("suggestions"):
            parts.append("alts: " + c["suggestions"])
        df.at[c["idx"], "conflict_note"] = " | ".join([p for p in parts if p])

    rows = []
    for c in conflict_rows:
        row = df.loc[c["idx"]]
        rows.append({
            "row_index": c["idx"],
            "course_id": row.get("course_id"),
            "title": row.get("title"),
            "day": c["day"],
            "time": f"{time_str(row.get('start_min'))}-{time_str(row.get('end_min'))}",
            "assigned_bldg": c["assigned_bldg"],
            "assigned_room": c["assigned_room"],
            "type": c["type"],
            "reason": c.get("reason",""),
            "other_course": c.get("other_course",""),
            "suggestions": c.get("suggestions","")
        })
    conflicts_df = pd.DataFrame(rows).drop_duplicates()
    return df, conflicts_df

def build_weekly_grid(df, room=None, slot_minutes=30):
    dff = df.copy()
    if room and room != "(All rooms)":
        spl = room.split("—") if "—" in room else room.split("-", 1)
        if len(spl) == 2:
            bldg, rm = spl
            dff = dff[(dff["assigned_bldg"].astype(str).str.strip() == bldg.strip()) &
                      (dff["assigned_room"].astype(str).str.strip() == rm.strip())]
    starts = dff["start_min"].dropna().tolist()
    ends = dff["end_min"].dropna().tolist()
    if not starts or not ends:
        day_start, day_end = 8*60, 18*60
    else:
        day_start = min(starts + [8*60])
        day_end = max(ends + [18*60])
    day_start = (day_start // slot_minutes) * slot_minutes
    day_end = int(math.ceil(day_end / slot_minutes) * slot_minutes)

    index = [time_str(t) for t in range(int(day_start), int(day_end), int(slot_minutes))]
    grid = pd.DataFrame("", index=index, columns=DAY_ORDER)
    from ast import literal_eval
    for _, row in dff.iterrows():
        label = f"{'🧪' if row.get('is_lab') else '📘'} {row.get('course_id')}"
        s, e = row.get("start_min"), row.get("end_min")
        if s is None or e is None:
            continue
        days_list = row.get("days_list", [])
        if isinstance(days_list, str):
            try: days_list = literal_eval(days_list)
            except Exception: days_list = []
        for d in days_list:
            for t in range(int((s // slot_minutes) * slot_minutes), int(math.ceil(e / slot_minutes) * slot_minutes), int(slot_minutes)):
                idx = time_str(t)
                if idx in grid.index and d in grid.columns:
                    if grid.at[idx, d] == "":
                        grid.at[idx, d] = label
                    else:
                        grid.at[idx, d] += " | ⚠️ " + label
    return grid

def build_beautiful_week_timeline(df, group_by="room"):
    records = []
    from ast import literal_eval
    for _, row in df.iterrows():
        s, e = row.get("start_min"), row.get("end_min")
        if s is None or e is None:
            continue
        b = norm_str(row.get("assigned_bldg")); r = norm_str(row.get("assigned_room"))
        room_label = f"{b}-{r}" if (b or r) else "(no room)"
        typ = "LAB" if row.get("is_lab") else "LEC"
        days_list = row.get("days_list", [])
        if isinstance(days_list, str):
            try: days_list = literal_eval(days_list)
            except Exception: days_list = []
        for d in days_list:
            start_dt = minutes_to_dt(d, s)
            end_dt = minutes_to_dt(d, e)
            records.append({
                "y": room_label if group_by == "room" else row.get("course_id"),
                "type": typ,
                "start": start_dt,
                "end": end_dt,
                "label": f"{typ} {row.get('course_id')} • {room_label}"
            })
    if not records:
        return None
    y_categories = sorted(set([r["y"] for r in records]))
    fig = go.Figure()
    color_map = {"LAB": "#4CAF50", "LEC": "#2196F3"}
    for rec in records:
        fig.add_trace(go.Bar(
            x=[(rec["end"]-rec["start"]).total_seconds()/3600.0],
            y=[rec["y"]],
            base=[rec["start"]],
            orientation="h",
            hovertext=rec["label"],
            marker_color=color_map.get(rec["type"], "#888")
        ))
    fig.update_layout(
        barmode="overlay",
        height=520,
        margin=dict(l=10,r=10,t=30,b=10),
        xaxis=dict(type="date", title="Time", tickformat="%a %H:%M"),
        yaxis=dict(categoryorder="array", categoryarray=y_categories, title= "Room" if group_by=="room" else "Course"),
        showlegend=False
    )
    return fig

# ----- Export helpers: modify the uploaded workbook in-place -----
def compute_changes(df_current, df_original, conflicts_df):
    orig_b = df_original.get("bldg")
    orig_r = df_original.get("room")
    cur_b = df_current.get("assigned_bldg")
    cur_r = df_current.get("assigned_room")
    reason_map = {}
    if conflicts_df is not None and not conflicts_df.empty and "row_index" in conflicts_df.columns:
        auto = conflicts_df[conflicts_df["type"] == "auto-resolved"]
        for _, row in auto.iterrows():
            reason_map[int(row["row_index"])] = row.get("reason", "Auto-resolved")
    changes = []
    for i in range(len(df_current)):
        ob = norm_str(orig_b.iloc[i]) if isinstance(orig_b, pd.Series) and i < len(orig_b) else ""
        oroom = norm_str(orig_r.iloc[i]) if isinstance(orig_r, pd.Series) and i < len(orig_r) else ""
        nb = norm_str(cur_b.iloc[i]) if isinstance(cur_b, pd.Series) and i < len(cur_b) else ""
        nr = norm_str(cur_r.iloc[i]) if isinstance(cur_r, pd.Series) and i < len(cur_r) else ""
        changed = (ob != "" or oroom != "") and (nb != ob or nr != oroom)
        if changed:
            reason = reason_map.get(i, "Manual change")
            change_text = f"{ob}-{oroom} → {nb}-{nr}"
        else:
            reason = ""
            change_text = ""
        changes.append((changed, reason, change_text))
    return changes

def export_inplace_modify_uploaded(upload_bytes, sheet_name, df_current, df_original, conflicts_df, final_filename):
    """
    Replaces the chosen sheet in the *uploaded* workbook with the modified table,
    adds Change_Reason and Change columns, and color-codes changed rows.
    Other sheets are preserved as-is.
    """
    # Build output DataFrame: keep df_current + reason/change
    changes = compute_changes(df_current, df_original, conflicts_df)
    changed_col = [c[0] for c in changes]
    reason_col = [c[1] for c in changes]
    change_txt = [c[2] for c in changes]
    out_df = df_current.copy()
    out_df["Change_Reason"] = reason_col
    out_df["Change"] = change_txt

    # Load workbook from the uploaded bytes
    bio = io.BytesIO(upload_bytes)
    wb = load_workbook(filename=bio)
    # Replace or create sheet
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name, index=idx)
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Write the DataFrame to the sheet
    # Header
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for j, col in enumerate(out_df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.fill = header_fill
        c.font = Font(bold=True)
    # Body
    for i, (_, row) in enumerate(out_df.iterrows(), start=2):
        for j, col in enumerate(out_df.columns, start=1):
            ws.cell(row=i, column=j, value=row[col])

    # Color-code changed room cells (both assigned_* and original bldg/room if present)
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cols = list(out_df.columns)
    col_map = {name: (cols.index(name) + 1) for name in cols}
    for i, was_changed in enumerate(changed_col, start=2):
        if was_changed:
            for target in ["assigned_bldg", "assigned_room", "bldg", "room"]:
                if target in col_map:
                    ws.cell(row=i, column=col_map[target]).fill = yellow

    # Wider columns for reason/change
    for target in ["Change_Reason", "Change"]:
        if target in col_map:
            ws.column_dimensions[chr(64 + col_map[target])].width = 35

    # Save to buffer
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, final_filename


def find_header_map(ws, header_row):
    """Return a dict: normalized header -> column index (1-based)."""
    def norm(h):
        return str(h).strip().lower().replace(" ", "_") if h is not None else ""
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        header_map[norm(val)] = col
    return header_map

def export_inplace_targeted(upload_bytes, sheet_name, header_row, df_current, df_original, conflicts_df, final_filename):
    """
    Modify ONLY BLDG/ROOM cells in the existing sheet and append Change_Reason/Change columns.
    Keep all other formatting, widths, and sheets intact.
    """
    # Compute change metadata
    changes = compute_changes(df_current, df_original, conflicts_df)
    reason_col = [c[1] for c in changes]
    change_txt = [c[2] for c in changes]

    bio = io.BytesIO(upload_bytes)
    wb = load_workbook(filename=bio)
    if sheet_name not in wb.sheetnames:
        # Fallback to replacing (shouldn't happen since we choose an existing sheet)
        return export_inplace_modify_uploaded(upload_bytes, sheet_name, df_current, df_original, conflicts_df, final_filename)

    ws = wb[sheet_name]

    # Map headers to columns using the provided header row (1-based)
    header_map = find_header_map(ws, header_row)
    # Candidate names
    bldg_col = header_map.get("bldg") or header_map.get("building") or header_map.get("assigned_bldg")
    room_col = header_map.get("room") or header_map.get("assigned_room")

    # Ensure Change_Reason and Change columns exist (append at end if missing)
    def ensure_col(name):
        norm = name.lower().replace(" ", "_")
        col_idx = header_map.get(norm)
        if not col_idx:
            col_idx = ws.max_column + 1
            ws.cell(row=header_row, column=col_idx, value=name)
            ws.cell(row=header_row, column=col_idx).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws.cell(row=header_row, column=col_idx).font = Font(bold=True)
            # update header_map
            header_map[norm] = col_idx
        return col_idx

    reason_idx = ensure_col("Change_Reason")
    change_idx = ensure_col("Change")

    # Highlight fill
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Data starts on the next row after header
    start_row = header_row + 1
    # Determine which df columns to pull
    cur_b = df_current.get("assigned_bldg", df_current.get("bldg"))
    cur_r = df_current.get("assigned_room", df_current.get("room"))

    # Iterate through dataframe rows and write back to sheet rows
    for i in range(len(df_current)):
        excel_row = start_row + i
        # Skip if row exceeds current sheet's used range; still safe to write (openpyxl will extend)
        # Write Change columns
        ws.cell(row=excel_row, column=reason_idx, value=reason_col[i] if i < len(reason_col) else "")
        ws.cell(row=excel_row, column=change_idx, value=change_txt[i] if i < len(change_txt) else "")

        # Update BLDG/ROOM cells only if those columns exist
        if bldg_col:
            new_b = cur_b.iloc[i] if i < len(cur_b) else None
            if new_b is not None and new_b != ws.cell(row=excel_row, column=bldg_col).value:
                ws.cell(row=excel_row, column=bldg_col, value=new_b)
                ws.cell(row=excel_row, column=bldg_col).fill = yellow
        if room_col:
            new_r = cur_r.iloc[i] if i < len(cur_r) else None
            if new_r is not None and new_r != ws.cell(row=excel_row, column=room_col).value:
                ws.cell(row=excel_row, column=room_col, value=new_r)
                ws.cell(row=excel_row, column=room_col).fill = yellow

    # Widen Change columns a bit
    try:
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(reason_idx)].width = 35
        ws.column_dimensions[get_column_letter(change_idx)].width = 35
    except Exception:
        pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, final_filename

# -------------------------------
# UI
# -------------------------------
st.title("🏫 Room Assignment & Conflict Resolver")
st.caption("Upload your Excel, use **Section_Act** (LAB/LEC) to prioritize labs, auto-resolve conflicts, visualize weekly planner & timeline, and export by modifying your uploaded workbook.")

with st.sidebar:
    st.header("1) Upload")
    up = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])
    sheet_choice = None
    header_row_choice = None
    st.markdown("---")
    st.header("2) Options")
    prioritize_labs = st.checkbox("Prioritize Labs over Lectures", value=True)
    slot_minutes = st.slider("Weekly planner slot size (minutes)", min_value=15, max_value=60, step=15, value=30)
    st.markdown("---")
    st.header("3) Actions")
    run_auto_btn = st.button("🔁 Run Auto-Assign & Detect Conflicts", use_container_width=True)
    export_btn = st.button("💾 Prepare Export", use_container_width=True, help="Downloads your original file name, with the chosen sheet replaced by the modified table (changes color-coded).")

# Session vars
if "df" not in st.session_state:
    st.session_state.df = None
if "conflicts" not in st.session_state:
    st.session_state.conflicts = pd.DataFrame()
if "original" not in st.session_state:
    st.session_state.original = None
if "lab_mode" not in st.session_state:
    st.session_state.lab_mode = "section_act"
if "upload_name" not in st.session_state:
    st.session_state.upload_name = None
if "upload_bytes" not in st.session_state:
    st.session_state.upload_bytes = None
if "sheet_name" not in st.session_state:
    st.session_state.sheet_name = "Assigned"

# Load & parse
if up is not None and st.session_state.df is None:
    try:
        st.session_state.upload_name = up.name
        st.session_state.upload_bytes = up.getvalue()  # Save original bytes for in-place export
        xls = pd.ExcelFile(up)
        sheets = xls.sheet_names
        st.success(f"Found sheets: {', '.join(sheets)}")
        sheet_choice = st.selectbox("Choose sheet", sheets, index=0, key="sheet_select")
        st.session_state.sheet_name = sheet_choice
        raw = pd.read_excel(xls, sheet_choice, header=None)
        hrow = detect_header_row(raw)
        st.info(f"Detected header row at index {hrow}. Change it below if needed.")
        header_row_choice = st.number_input("Header row index", min_value=0, max_value=int(max(0, len(raw)-1)), value=int(hrow), step=1, key="hdr_row")
        st.session_state.header_row = int(header_row_choice) + 1  # 1-based for openpyxl
        df = pd.read_excel(xls, sheet_choice, header=header_row_choice)
        df = df.dropna(how="all")
        df = normalize_columns(df)

        required = ["subject","number","title","sec","days","start","end"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            st.warning(f"Some expected columns are missing: {missing}. The app will still try to proceed.")
        df["course_id"] = (df.get("subject", "").astype(str) + "-" + df.get("number","").astype(str) + "-" + df.get("sec","").astype(str))
        df["days_list"] = df.get("days", "").apply(parse_days)
        df["start_min"] = df.get("start").apply(parse_time)
        df["end_min"] = df.get("end").apply(parse_time)
        if "assigned_bldg" not in df.columns:
            df["assigned_bldg"] = df.get("bldg")
        if "assigned_room" not in df.columns:
            df["assigned_room"] = df.get("room")

        # LAB/LEC classification
        sec_present = "section_act" in df.columns
        if sec_present:
            sec_vals = df["section_act"].astype(str).str.upper().str.strip()
            if not sec_vals.replace("NAN","").replace("","MISSING").ne("MISSING").any():
                st.warning("`Section_Act` column is present but empty. Choose how to determine Labs:")
                fallback_choice = st.radio("LAB detection mode", options=["Ask me later (treat all as LEC)","Detect by Title contains 'lab'"], index=0)
                lab_mode = "lecture" if "Ask" in fallback_choice else "title"
            else:
                lab_mode = "section_act"
        else:
            st.warning("No `Section_Act` column found. Choose how to determine Labs:")
            fallback_choice = st.radio("LAB detection mode", options=["Detect by Title contains 'lab'","Treat all as LEC"], index=0)
            lab_mode = "title" if "Title" in fallback_choice else "lecture"

        st.session_state.lab_mode = lab_mode
        if lab_mode == "section_act":
            df["is_lab"] = df["section_act"].astype(str).str.upper().str.strip().eq("LAB")
        elif lab_mode == "title":
            df["is_lab"] = df.get("title","").astype(str).str.lower().str.contains("lab")
        else:
            df["is_lab"] = False

        st.session_state.df = df.copy()
        st.session_state.original = df.copy()
        st.success("Parsed! Scroll to preview and next steps.")
    except Exception as e:
        st.error(f"Failed to parse file: {e}")

if st.session_state.df is not None:
    df = st.session_state.df

    # Preview
    st.subheader("Preview")
    c1, c2, c3, c4 = st.columns([2,2,1,1])
    with c1: st.metric("Rows", len(df))
    with c2: st.metric("Labs (current mode)", int(df["is_lab"].sum()))
    with c3: st.metric("With rooms", int((df["assigned_bldg"].notna() & df["assigned_room"].notna()).sum()))
    with c4: st.metric("Missing rooms", int((df["assigned_bldg"].isna() | df["assigned_room"].isna()).sum()))
    st.dataframe(df.head(40), use_container_width=True)

    # Run auto-assign & detect conflicts
    if run_auto_btn:
        df2, conflicts = run_auto_resolve(df.copy(), prioritize_labs=prioritize_labs, max_rounds=3)
        st.session_state.df = df2
        st.session_state.conflicts = conflicts
        st.success("Auto-assign & conflict detection complete.")

    # Conflicts panel
    st.subheader("Conflicts & Suggestions")
    conflicts = st.session_state.conflicts
    if conflicts is not None and not conflicts.empty:
        st.write("Auto-assign prioritizes Labs. Use the tools below to resolve or move entries.")

        st.dataframe(conflicts, use_container_width=True, height=260)

        rooms = build_rooms(st.session_state.df)
        if not rooms:
            st.info("No known rooms found in the data. Make sure bldg/room columns exist or some rows have rooms assigned.")

        idx_options = conflicts["row_index"].tolist()
        pick_idx = st.selectbox("Pick a conflicted row to modify", idx_options)

        row = st.session_state.df.loc[pick_idx]
        st.write(f"**{row.get('course_id')}** — {row.get('title')} | Days: {''.join(row.get('days_list', []))} | Time: {time_str(row.get('start_min'))}-{time_str(row.get('end_min'))}")
        cur_b, cur_r = norm_str(row.get("assigned_bldg")), norm_str(row.get("assigned_room"))
        st.write(f"Current room: **{cur_b}-{cur_r}**")

        st.markdown("**Room search options**")
        same_time_only = st.checkbox("Only show rooms free at the **same time**", value=True, help="Untick to search for flexible time slots with the same duration.")

        days = row.get("days_list", [])
        choose_day = st.selectbox("Choose day to search", options=days if days else DAY_ORDER)

        occupancy = rebuild_occupancy(st.session_state.df)
        if same_time_only:
            alts = list_free_rooms_same_time(occupancy, st.session_state.df, rooms, choose_day, row.get("start_min"), row.get("end_min"),
                                             limit=12, exclude={(cur_b, cur_r)})
            alt_label = [f"{b}-{r}" for (b,r) in alts] if alts else ["(no same-time alternatives)"]
            chosen = st.selectbox("Alternative room (same time)", options=alt_label)
            if st.button("Apply room change", type="primary"):
                if alts:
                    nb, nr = alts[alt_label.index(chosen)]
                    st.session_state.df.at[pick_idx, "assigned_bldg"] = nb
                    st.session_state.df.at[pick_idx, "assigned_room"] = nr
                    st.success(f"Updated to {nb}-{nr}. Re-run Auto-Assign to recompute conflicts.")
                else:
                    st.warning("No alternatives available at the same time.")
        else:
            duration = int(row.get("end_min") - row.get("start_min")) if pd.notna(row.get("start_min")) and pd.notna(row.get("end_min")) else 90
            st.caption(f"Searching flexible slots of duration ~{duration} minutes on {choose_day}.")
            suggestions = list_free_slots_across_rooms(st.session_state.df, choose_day, duration, day_start=8*60, day_end=20*60, limit=25, exclude_room=(cur_b, cur_r))
            if suggestions:
                opt_list = [f"{x['room']} • {x['day']} • {time_str(x['start'])}-{time_str(x['end'])}" for x in suggestions]
                st.dataframe(pd.DataFrame([{
                    "Room": s["room"],
                    "Day": s["day"],
                    "Start": time_str(s["start"]),
                    "End": time_str(s["end"])
                } for s in suggestions]), use_container_width=True, height=280)
                pick = st.selectbox("Pick an alternative slot", options=opt_list)
                if st.button("Apply room + time change", type="primary"):
                    sel = suggestions[opt_list.index(pick)]
                    nb, nr = sel["room"].split("-", 1)
                    st.session_state.df.at[pick_idx, "assigned_bldg"] = nb
                    st.session_state.df.at[pick_idx, "assigned_room"] = nr
                    st.session_state.df.at[pick_idx, "start_min"] = sel["start"]
                    st.session_state.df.at[pick_idx, "end_min"] = sel["end"]
                    st.session_state.df.at[pick_idx, "days_list"] = [sel["day"]]
                    if "start" in st.session_state.df.columns:
                        st.session_state.df.at[pick_idx, "start"] = int(time_str(sel["start"]).replace(":", ""))
                    if "end" in st.session_state.df.columns:
                        st.session_state.df.at[pick_idx, "end"] = int(time_str(sel["end"]).replace(":", ""))
                    st.success(f"Moved to {sel['room']} on {sel['day']} {time_str(sel['start'])}-{time_str(sel['end'])}. Re-run Auto-Assign to recompute conflicts.")
            else:
                st.info("No flexible slots found for that duration within 08:00–20:00.")

    else:
        st.info("No conflicts detected yet or you haven't run auto-assign.")

    # Weekly Planner (Grid)
    st.subheader("Weekly Planner (Grid)")
    rooms_list = build_rooms(st.session_state.df)
    room_labels = [f"{b}-{r}" for (b,r) in rooms_list]
    room_labels_vis = [lbl.replace("-", "—", 1) for lbl in room_labels]
    room_choice = st.selectbox("Filter by room (optional)", options=["(All rooms)"] + room_labels_vis, index=0, key="room_filter_grid")
    selected_room = None if room_choice == "(All rooms)" else room_choice
    grid = build_weekly_grid(st.session_state.df, room=selected_room, slot_minutes=slot_minutes)
    st.write("Grid legend: 🧪 Lab • 📘 Lecture • ⚠️ Overlap in same cell")
    st.dataframe(grid, use_container_width=True, height=400)

    # Weekly Timeline
    st.subheader("Weekly Timeline (Interactive)")
    group_by = st.radio("Group by", options=["room", "course"], horizontal=True, index=0, key="group_by_timeline")
    fig = build_beautiful_week_timeline(st.session_state.df, group_by=group_by)
    if fig:
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.write("Nothing to plot.")

    # Export (in-place modification)
    if export_btn:
        if not st.session_state.upload_bytes:
            st.error("Missing original upload bytes; please re-upload the Excel file.")
        else:
            buffer, fname = export_inplace_targeted(
                upload_bytes=st.session_state.upload_bytes,
                sheet_name=st.session_state.sheet_name or "Assigned",
                header_row=st.session_state.get('header_row', 1),
                df_current=st.session_state.df,
                df_original=st.session_state.original,
                conflicts_df=st.session_state.conflicts,
                final_filename=st.session_state.upload_name or "assigned_schedule.xlsx"
            )
            st.download_button("⬇️ Download Modified Workbook", data=buffer, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Reset
    if st.button("Reset to uploaded data"):
        st.session_state.df = st.session_state.original.copy()
        st.session_state.conflicts = pd.DataFrame()
        st.success("Reset complete.")
